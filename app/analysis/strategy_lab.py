"""
strategy_lab.py — Redline Quant Strategy Lab Analysis Engine

Refactored from stock_analyzer.py.
Core logic is unchanged (period="60d", interval="15m", SMA20, SMA50,
daily/weekly aggregations, openpyxl charts).

No code executes at import time — all analysis is triggered via run_strategy_analysis().
"""

from __future__ import annotations

import os
import uuid
from pathlib import Path
from typing import Optional

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference

# Root output directory — relative to this file's package root
_DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "jobs"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _ensure_jobs_dir(job_id: str) -> Path:
    job_path = _DATA_DIR / job_id
    job_path.mkdir(parents=True, exist_ok=True)
    return job_path


def _safe_float(val) -> Optional[float]:
    """Convert a pandas scalar to float safely, returning None for NaN/inf."""
    try:
        f = float(val)
        if pd.isna(f) or f != f:  # NaN check
            return None
        return round(f, 2)
    except Exception:
        return None


def _format_date(ts) -> str:
    try:
        return pd.Timestamp(ts).strftime("%d %b %Y")
    except Exception:
        return str(ts)


def _clean_company_name(backend_symbol: str, provided_name: str) -> str:
    """
    Use the provided company name if non-empty, otherwise derive from ticker.
    Never just strip .NS and call it a company.
    """
    KNOWN = {
        "ITC": "ITC Limited",
        "RELIANCE": "Reliance Industries Ltd.",
        "TCS": "Tata Consultancy Services",
        "INFY": "Infosys Limited",
        "HDFCBANK": "HDFC Bank Ltd.",
        "ICICIBANK": "ICICI Bank Ltd.",
        "SBIN": "State Bank of India",
        "BHARTIARTL": "Bharti Airtel Ltd.",
        "LT": "Larsen & Toubro Ltd.",
        "MARUTI": "Maruti Suzuki India Ltd.",
        "TATAMOTORS": "Tata Motors Ltd.",
        "AXISBANK": "Axis Bank Ltd.",
        "KOTAKBANK": "Kotak Mahindra Bank Ltd.",
        "HINDUNILVR": "Hindustan Unilever Ltd.",
        "AAPL": "Apple Inc.",
        "MSFT": "Microsoft Corporation",
        "GOOGL": "Alphabet Inc.",
        "AMZN": "Amazon.com Inc.",
        "NVDA": "NVIDIA Corporation",
        "META": "Meta Platforms Inc.",
        "TSLA": "Tesla Inc.",
        "NFLX": "Netflix Inc.",
        "AMD": "Advanced Micro Devices",
        "INTC": "Intel Corporation",
        "JPM": "JPMorgan Chase & Co.",
        "V": "Visa Inc.",
        "WMT": "Walmart Inc.",
        "DIS": "The Walt Disney Company",
    }
    if provided_name and provided_name.strip():
        return provided_name.strip()
    clean = backend_symbol.upper().replace(".NS", "").replace(".BO", "")
    return KNOWN.get(clean, clean)


# ---------------------------------------------------------------------------
# 1. DATA DOWNLOAD (exact same parameters as stock_analyzer.py)
# ---------------------------------------------------------------------------

def _download_data(backend_symbol: str) -> pd.DataFrame:
    """Download 15-minute intraday data for the last 60 days."""
    data = yf.download(
        backend_symbol,
        period="60d",
        interval="15m",
        auto_adjust=False,
        progress=False,
    )
    if data.empty:
        raise ValueError(f"No 15-minute market data was found for '{backend_symbol}'.")

    # Fix MultiIndex columns (yfinance returns them for single ticker sometimes)
    if isinstance(data.columns, pd.MultiIndex):
        data.columns = data.columns.droplevel(1)

    # Strip timezone for Excel compatibility
    if data.index.tz is not None:
        data.index = data.index.tz_localize(None)

    data = data.dropna()

    if len(data) < 50:
        raise ValueError(
            f"Not enough intraday data to calculate indicators for '{backend_symbol}'. "
            "Only the last 60 days of 15-minute data are available from Yahoo Finance."
        )

    return data


# ---------------------------------------------------------------------------
# 2. INDICATOR CALCULATION (identical to stock_analyzer.py)
# ---------------------------------------------------------------------------

def _calculate_indicators(data: pd.DataFrame) -> pd.DataFrame:
    data = data.copy()

    # 15-minute period return
    data["15m_Return_%"] = data["Close"].pct_change() * 100

    # SMA20 and SMA50
    data["SMA20"] = data["Close"].rolling(20).mean()
    data["SMA50"] = data["Close"].rolling(50).mean()

    return data


# ---------------------------------------------------------------------------
# 3. DAILY AGGREGATION (identical to stock_analyzer.py)
# ---------------------------------------------------------------------------

def _calc_daily(data: pd.DataFrame) -> pd.DataFrame:
    daily = data.resample("D").agg({
        "Open": "first",
        "High": "max",
        "Low": "min",
        "Close": ["mean", "last"],
        "Volume": "sum",
    })
    daily = daily.dropna()
    daily.columns = [
        "Open",
        "Highest_Price",
        "Lowest_Price",
        "Average_Close",
        "Day_End_Close",
        "Total_Volume",
    ]
    daily["Daily_Return_%"] = daily["Day_End_Close"].pct_change() * 100
    return daily


# ---------------------------------------------------------------------------
# 4. WEEKLY AGGREGATION (identical to stock_analyzer.py)
# ---------------------------------------------------------------------------

def _calc_weekly(data: pd.DataFrame) -> pd.DataFrame:
    weekly = data.resample("W").agg({
        "Open": "first",
        "High": "max",
        "Low": "min",
        "Close": ["mean", "last"],
        "Volume": "sum",
    })
    weekly = weekly.dropna()
    weekly.columns = [
        "Open",
        "Highest_Price",
        "Lowest_Price",
        "Average_Close",
        "Week_End_Close",
        "Total_Volume",
    ]
    weekly["Weekly_Return_%"] = weekly["Week_End_Close"].pct_change() * 100
    return weekly


# ---------------------------------------------------------------------------
# 5. CSV EXPORT
# ---------------------------------------------------------------------------

def create_csv_report(data: pd.DataFrame, clean_symbol: str, job_path: Path) -> Path:
    """Save processed 15-minute data as CSV. Returns the file path."""
    csv_path = job_path / f"{clean_symbol}_15m.csv"
    data.to_csv(csv_path)
    return csv_path


# ---------------------------------------------------------------------------
# 6. EXCEL EXPORT WITH CHARTS (identical logic to stock_analyzer.py)
# ---------------------------------------------------------------------------

def create_excel_report(
    data: pd.DataFrame,
    daily: pd.DataFrame,
    weekly: pd.DataFrame,
    clean_symbol: str,
    company: str,
    job_path: Path,
) -> Path:
    """
    Save Excel workbook with 3 sheets + openpyxl charts.
    Preserves exact chart structure from the original stock_analyzer.py.
    """
    excel_path = job_path / f"{clean_symbol}_15m_Analysis.xlsx"

    with pd.ExcelWriter(str(excel_path), engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="15m Data")
        daily.to_excel(writer, sheet_name="Daily Analysis")
        weekly.to_excel(writer, sheet_name="Weekly Analysis")

    # Re-open to add charts (same as stock_analyzer.py sections 16-20)
    workbook = load_workbook(str(excel_path))

    # ----- Chart 1: Daily Average Price (LineChart) -----
    sheet = workbook["Daily Analysis"]

    chart1 = LineChart()
    chart1.title = f"{company} - Daily Average Price"
    chart1.y_axis.title = "Price"
    chart1.x_axis.title = "Date"

    data_ref = Reference(sheet, min_col=5, min_row=1, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(categories)
    chart1.height = 10
    chart1.width = 20

    sheet.add_chart(chart1, "J2")

    # ----- Chart 2: Daily Returns (BarChart) -----
    chart2 = BarChart()
    chart2.title = f"{company} - Daily Returns"
    chart2.y_axis.title = "Return (%)"
    chart2.x_axis.title = "Date"

    data_ref2 = Reference(sheet, min_col=8, min_row=1, max_row=sheet.max_row)

    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(categories)
    chart2.height = 10
    chart2.width = 20

    sheet.add_chart(chart2, "J22")

    # ----- Chart 3: Weekly Returns (BarChart) -----
    sheet2 = workbook["Weekly Analysis"]

    chart3 = BarChart()
    chart3.title = f"{company} - Weekly Returns"
    chart3.y_axis.title = "Return (%)"
    chart3.x_axis.title = "Week"

    data_ref3 = Reference(sheet2, min_col=8, min_row=1, max_row=sheet2.max_row)
    categories2 = Reference(sheet2, min_col=1, min_row=2, max_row=sheet2.max_row)

    chart3.add_data(data_ref3, titles_from_data=True)
    chart3.set_categories(categories2)
    chart3.height = 10
    chart3.width = 20

    sheet2.add_chart(chart3, "J2")

    workbook.save(str(excel_path))

    return excel_path


# ---------------------------------------------------------------------------
# 7. MAIN ENTRY POINT
# ---------------------------------------------------------------------------

def run_strategy_analysis(
    backend_symbol: str,
    company_name: str = "",
    currency: str = "USD",
) -> dict:
    """
    Execute the full 60-day intraday analysis pipeline.

    Args:
        backend_symbol: yfinance ticker (e.g. "ITC.NS" or "AAPL")
        company_name: Human-readable company name (e.g. "ITC Limited")
        currency: "INR" or "USD"

    Returns:
        Structured dict with summary metrics, series data, and download paths.
    """
    # Derive clean display symbol
    clean_symbol = backend_symbol.upper().replace(".NS", "").replace(".BO", "")
    company = _clean_company_name(backend_symbol, company_name)

    # --- Download ---
    data = _download_data(backend_symbol)

    # --- Indicators ---
    data = _calculate_indicators(data)

    # --- Aggregations ---
    daily = _calc_daily(data)
    weekly = _calc_weekly(data)

    # --- Best/Worst ---
    daily_valid = daily.dropna(subset=["Daily_Return_%"])
    weekly_valid = weekly.dropna(subset=["Weekly_Return_%"])

    best_day_idx = daily_valid["Daily_Return_%"].idxmax()
    worst_day_idx = daily_valid["Daily_Return_%"].idxmin()
    best_week_idx = weekly_valid["Weekly_Return_%"].idxmax()
    worst_week_idx = weekly_valid["Weekly_Return_%"].idxmin()

    # --- Price extremes ---
    highest_price = _safe_float(data["High"].max())
    lowest_price = _safe_float(data["Low"].min())

    # --- Overall return ---
    overall_return = _safe_float(
        (data["Close"].iloc[-1] / data["Close"].iloc[0] - 1) * 100
    )

    # --- Last SMA values ---
    last_sma20 = _safe_float(data["SMA20"].dropna().iloc[-1]) if not data["SMA20"].dropna().empty else None
    last_sma50 = _safe_float(data["SMA50"].dropna().iloc[-1]) if not data["SMA50"].dropna().empty else None

    # --- Series for frontend charts ---
    timestamps = [t.strftime("%Y-%m-%d %H:%M") for t in data.index]
    close_series = [_safe_float(v) for v in data["Close"]]
    sma20_series = [_safe_float(v) for v in data["SMA20"]]
    sma50_series = [_safe_float(v) for v in data["SMA50"]]
    return_series = [_safe_float(v) for v in data["15m_Return_%"]]

    daily_dates = [_format_date(t) for t in daily_valid.index]
    daily_returns = [_safe_float(v) for v in daily_valid["Daily_Return_%"]]

    weekly_dates = [_format_date(t) for t in weekly_valid.index]
    weekly_returns = [_safe_float(v) for v in weekly_valid["Weekly_Return_%"]]

    # --- Generate files ---
    job_id = str(uuid.uuid4())
    job_path = _ensure_jobs_dir(job_id)

    csv_path = create_csv_report(data, clean_symbol, job_path)
    excel_path = create_excel_report(data, daily, weekly, clean_symbol, company, job_path)

    return {
        "success": True,
        "job_id": job_id,
        "symbol": clean_symbol,
        "company": company,
        "currency": currency,
        "period": "60d",
        "interval": "15m",
        "data_points": len(data),

        "summary": {
            "highestPrice": highest_price,
            "lowestPrice": lowest_price,
            "overallReturn": overall_return,
            "lastSMA20": last_sma20,
            "lastSMA50": last_sma50,
        },

        "bestDay": {
            "date": _format_date(best_day_idx),
            "return": _safe_float(daily_valid.loc[best_day_idx, "Daily_Return_%"]),
        },
        "worstDay": {
            "date": _format_date(worst_day_idx),
            "return": _safe_float(daily_valid.loc[worst_day_idx, "Daily_Return_%"]),
        },
        "bestWeek": {
            "date": _format_date(best_week_idx),
            "return": _safe_float(weekly_valid.loc[best_week_idx, "Weekly_Return_%"]),
        },
        "worstWeek": {
            "date": _format_date(worst_week_idx),
            "return": _safe_float(weekly_valid.loc[worst_week_idx, "Weekly_Return_%"]),
        },

        "series": {
            "timestamps": timestamps,
            "close": close_series,
            "sma20": sma20_series,
            "sma50": sma50_series,
            "returns": return_series,
            "dailyDates": daily_dates,
            "dailyReturns": daily_returns,
            "weeklyDates": weekly_dates,
            "weeklyReturns": weekly_returns,
        },

        "download": {
            "csv": f"/api/v1/strategy-lab/download/{job_id}/csv",
            "excel": f"/api/v1/strategy-lab/download/{job_id}/excel",
            "csvFilename": f"{clean_symbol}_15m.csv",
            "excelFilename": f"{clean_symbol}_15m_Analysis.xlsx",
        },

        # Internal use by download endpoints — not exposed to frontend
        "_csv_path": str(csv_path),
        "_excel_path": str(excel_path),
    }
