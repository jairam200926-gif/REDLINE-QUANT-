import yfinance as yf
import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference


# ============================================================
# 1. STOCK NAME
# ============================================================

stock = "NVDA"


# ============================================================
# 2. DOWNLOAD STOCK DATA (15-MINUTE INTERVAL)
# ============================================================

print("Downloading 15-minute data...")

data = yf.download(
    stock,
    period="60d",  
    interval="15m", 
    auto_adjust=False
)

if data.empty:
    print("ERROR: No data found.")
    print("Check the stock symbol.")
    exit()


# ============================================================
# 3. FIX YFINANCE COLUMNS & TIMEZONE
# ============================================================

if isinstance(data.columns, pd.MultiIndex):
    data.columns = data.columns.droplevel(1)

# Strip timezone metadata for Excel compatibility
if data.index.tz is not None:
    data.index = data.index.tz_localize(None)

data = data.dropna()


# ============================================================
# 4. 15-MINUTE PERIOD RETURN
# ============================================================

data["15m_Return_%"] = (
    data["Close"].pct_change() * 100
)


# ============================================================
# 5. 20-PERIOD AND 50-PERIOD MOVING AVERAGE
# ============================================================

data["SMA20"] = (
    data["Close"].rolling(20).mean()
)

data["SMA50"] = (
    data["Close"].rolling(50).mean()
)


# ============================================================
# 6. DAILY ANALYSIS
# ============================================================

daily = data.resample("D").agg({
    "Open": "first",
    "High": "max",
    "Low": "min",
    "Close": ["mean", "last"],
    "Volume": "sum"
})

daily = daily.dropna()

daily.columns = [
    "Open",
    "Highest_Price",
    "Lowest_Price",
    "Average_Close",
    "Day_End_Close",
    "Total_Volume"
]

daily["Daily_Return_%"] = (
    daily["Day_End_Close"]
    .pct_change()
    * 100
)


# ============================================================
# 7. WEEKLY ANALYSIS
# ============================================================

weekly = data.resample("W").agg({
    "Open": "first",
    "High": "max",
    "Low": "min",
    "Close": ["mean", "last"],
    "Volume": "sum"
})

weekly = weekly.dropna()

weekly.columns = [
    "Open",
    "Highest_Price",
    "Lowest_Price",
    "Average_Close",
    "Week_End_Close",
    "Total_Volume"
]

weekly["Weekly_Return_%"] = (
    weekly["Week_End_Close"]
    .pct_change()
    * 100
)


# ============================================================
# 8. BEST AND WORST PERFORMANCE
# ============================================================

daily_valid = daily.dropna(subset=["Daily_Return_%"])
weekly_valid = weekly.dropna(subset=["Weekly_Return_%"])

best_day = daily_valid["Daily_Return_%"].idxmax()
worst_day = daily_valid["Daily_Return_%"].idxmin()

best_week = weekly_valid["Weekly_Return_%"].idxmax()
worst_week = weekly_valid["Weekly_Return_%"].idxmin()


# ============================================================
# 9. HIGHEST / LOWEST PRICE
# ============================================================

highest_price = data["High"].max()
lowest_price = data["Low"].min()


# ============================================================
# 10. OVERALL RETURN
# ============================================================

overall_return = (
    (
        data["Close"].iloc[-1]
        /
        data["Close"].iloc[0]
    )
    - 1
) * 100


# ============================================================
# 11. COMPANY NAME
# ============================================================

company = stock.replace(".NS", "")


# ============================================================
# 12. PRINT SUMMARY
# ============================================================

print()
print("==========================================")
print("        60-DAY INTRADAY ANALYSIS SUMMARY")
print("==========================================")

print(f"Stock: {company}")
print()
print(f"Highest Price: ${highest_price:,.2f}")
print(f"Lowest Price: ${lowest_price:,.2f}")
print()
print(f"Best Day: {best_day.strftime('%B %d, %Y')}")
print(f"Best Daily Return: {daily.loc[best_day, 'Daily_Return_%']:.2f}%")
print()
print(f"Worst Day: {worst_day.strftime('%B %d, %Y')}")
print(f"Worst Daily Return: {daily.loc[worst_day, 'Daily_Return_%']:.2f}%")
print()
print(f"Best Week Ending: {best_week.strftime('%B %d, %Y')}")
print(f"Best Weekly Return: {weekly.loc[best_week, 'Weekly_Return_%']:.2f}%")
print()
print(f"Worst Week Ending: {worst_week.strftime('%B %d, %Y')}")
print(f"Worst Weekly Return: {weekly.loc[worst_week, 'Weekly_Return_%']:.2f}%")
print()
print(f"Overall 60-Day Return: {overall_return:.2f}%")


# ============================================================
# 13. CREATE DATA FOLDER
# ============================================================

os.makedirs("data", exist_ok=True)


# ============================================================
# 14. SAVE RAW DATA AS CSV
# ============================================================

csv_file = f"data/{company}_15m.csv"
data.to_csv(csv_file)


# ============================================================
# 15. CREATE EXCEL FILE
# ============================================================

excel_file = f"data/{company}_15m_Analysis.xlsx"

with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
    data.to_excel(writer, sheet_name="15m Data")
    daily.to_excel(writer, sheet_name="Daily Analysis")
    weekly.to_excel(writer, sheet_name="Weekly Analysis")


# ============================================================
# 16. OPEN EXCEL FILE
# ============================================================

workbook = load_workbook(excel_file)


# ============================================================
# 17. DAILY AVERAGE PRICE GRAPH
# ============================================================

sheet = workbook["Daily Analysis"]

chart1 = LineChart()
chart1.title = f"{company} - Daily Average Price"
chart1.y_axis.title = "Price ($)"
chart1.x_axis.title = "Date"

data_ref = Reference(sheet, min_col=5, min_row=1, max_row=sheet.max_row)
categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(categories)
chart1.height = 10
chart1.width = 20

sheet.add_chart(chart1, "J2")


# ============================================================
# 18. DAILY RETURN GRAPH
# ============================================================

chart2 = BarChart()
chart2.title = f"{company} - Daily Returns"
chart2.y_axis.title = "Return (%)"
chart2.x_axis.title = "Date"

data_ref = Reference(sheet, min_col=8, min_row=1, max_row=sheet.max_row)

chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(categories)
chart2.height = 10
chart2.width = 20

sheet.add_chart(chart2, "J22")


# ============================================================
# 19. WEEKLY RETURN GRAPH
# ============================================================

sheet2 = workbook["Weekly Analysis"]

chart3 = BarChart()
chart3.title = f"{company} - Weekly Returns"
chart3.y_axis.title = "Return (%)"
chart3.x_axis.title = "Week"

data_ref = Reference(sheet2, min_col=8, min_row=1, max_row=sheet2.max_row)
categories2 = Reference(sheet2, min_col=1, min_row=2, max_row=sheet2.max_row)

chart3.add_data(data_ref, titles_from_data=True)
chart3.set_categories(categories2)
chart3.height = 10
chart3.width = 20

sheet2.add_chart(chart3, "J2")


# ============================================================
# 20. SAVE EXCEL
# ============================================================

workbook.save(excel_file)


# ============================================================
# 21. FINISHED
# ============================================================

print()
print("==========================================")
print("       ANALYSIS COMPLETED SUCCESSFULLY")
print("==========================================")
print()
print(f"CSV saved at: {csv_file}")
print(f"Excel saved at: {excel_file}")